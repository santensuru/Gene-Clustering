__author__ = 'Djuned Fernando Djusdek'

from bs4 import BeautifulSoup
import urllib2
from openpyxl import load_workbook

wb = load_workbook('Gene_function.xlsx')
wb2 = wb.get_sheet_names()
ws = wb.get_sheet_by_name(wb2[0])

iter = 1

while True:
    if iter == 1214:
        wb.save('Gene_function2.xlsx')
        break
    
    gene_name = ws.cell(row = iter, column = 1).value
    print gene_name
    
    url = "http://www.genecards.org/cgi-bin/carddisp.pl?gene=" + gene_name
    
    hdr = {'Host': 'www.genecards.org',
       'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; WOW64; rv:39.0) Gecko/20100101 Firefox/39.0',
       'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8',
       'Accept-Language': 'en-US,en;q=0.5',
       'Accept-Encoding': 'gzip, deflate',
       'Cookie': 'visid_incap_146342=xCq2mxzgS3eijZ+20y3kKN/yblYAAAAAQkIPAAAAAACAUMhwAX2khpNomHubn/LYdgrv2R9hAP7p; incap_ses_218_146342=ijHxbNoIB1mBjgGKrn4GA9/yblYAAAAAvbGo8CMnbb4WX3iSTUy+lA==; nlbi_146342=RYbtfDXjbgp/Iw2nFHlFeAAAAABv4q31JmWPDntSktYeisHR',
       'Connection': 'keep-alive',
       'Cache-Control': 'max-age=0'}
    
    req = urllib2.Request(url, headers=hdr)
    
    try:
        page = urllib2.urlopen(req)

        soup = BeautifulSoup(page.read(), "html.parser")
        
        raw = soup.find("section", {"id": "_function"})

        try:
            raw = raw.find("tbody", {}).find_all("tr", {})
            
            genes_ontology = "";
            
            for i in range(len(raw)):
                genes_ontology += raw[i].find("a", {}).text + ";"
            
            ws.cell(row = iter, column = 2).value = genes_ontology
        except:
            ws.cell(row = iter, column = 2).value = "null"
            
    except urllib2.HTTPError, e:
        print e.fp.read()
    
    iter += 1
