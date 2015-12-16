__author__ = 'Djuned Fernando Djusdek'

from bs4 import BeautifulSoup
import urllib2
from openpyxl import load_workbook

wb = load_workbook('Gene_ontology.xlsx')
wb2 = wb.get_sheet_names()
ws = wb.get_sheet_by_name(wb2[0])

iter = 1

while True:
    if iter == 5:
        wb.save('Gene_ontology2.xlsx')
        break
    
    gene_ontology_id = ws.cell(row = iter, column = 1).value
    print gene_ontology_id
    
    url = "http://amigo.geneontology.org/amigo/term/" + gene_ontology_id
    
    hdr = {'Host': 'amigo.geneontology.org',
       'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; WOW64; rv:39.0) Gecko/20100101 Firefox/39.0',
       'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8',
       'Accept-Language': 'en-US,en;q=0.5',
       'Accept-Encoding': 'gzip, deflate',
       'Cookie': '_ga=GA1.2.1703632139.1450109154; __utma=243921171.1703632139.1450109154.1450254477.1450254477.1; __utmz=243921171.1450254477.1.1.utmcsr=google|utmccn=(organic)|utmcmd=organic|utmctr=(not%20provided); _gat=1',
       'Connection': 'keep-alive',
       'Cache-Control': 'max-age=0'}
    
    req = urllib2.Request(url, headers=hdr)
    
    try:
        page = urllib2.urlopen(req)

        soup = BeautifulSoup(page.read(), "html.parser")
        
        raw = soup.find("dl", {"class": "amigo-detail-info"})

        start = str(raw).find("<dt>Ontology</dt>")
        end = str(raw).find("<dt>Synonyms</dt>")
        contain = str(raw)[start:end]

        start = contain.find("<dd>") + 4
        end = contain.find("</dd>")

        ontology = contain[start:end]
        
        ws.cell(row = iter, column = 2).value = ontology
            
    except urllib2.HTTPError, e:
        print e.fp.read()
    
    iter += 1
